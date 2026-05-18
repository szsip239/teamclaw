"""
Excel ingestion pipeline (row → policy → chunks with hybrid index).

Ported from llm-rag/excel_qa.py, with SQLite swapped for the asyncpg
storage layer. The shape of the data is preserved: each Excel row becomes
a "policy" with structured metadata; the content field is chunked with
fixed-overlap splits; both rows and chunks live in Postgres with FTS
tokens + dense vectors.
"""

from __future__ import annotations

import logging
import re
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.config import RequestCredentials
from app.embedding import encode_texts
from app.fts import tokenize_for_index
from app.storage import (
    list_excel_chunks_for_embedding,
    replace_excel_policies,
    save_excel_chunk_vector,
    save_excel_config,
)

logger = logging.getLogger(__name__)

DEFAULT_EXCEL_CHUNK_SIZE = 1000
DEFAULT_EXCEL_CHUNK_OVERLAP = 100
KNOWN_HEADER_WORDS = {
    "序号", "标题", "级别", "政策发文号", "政策发文字号", "政策原文", "正文", "内容",
    "名称", "类型", "地区", "分类", "来源", "title", "content", "level",
}
DEFAULT_IGNORE_FIELDS = {"序号", "编号", "id", "ID"}


# ── Excel reading ────────────────────────────────────────────────────

def _cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _is_xls_source(path_or_bytes: str | Path | bytes) -> bool:
    if isinstance(path_or_bytes, bytes):
        return path_or_bytes.startswith(b"\xd0\xcf\x11\xe0")
    return Path(path_or_bytes).suffix.lower() == ".xls"


def _read_excel_rows(path_or_bytes: str | Path | bytes) -> tuple[str, list[list[str]]]:
    if _is_xls_source(path_or_bytes):
        try:
            import xlrd
        except ImportError as exc:
            raise ValueError(
                "Reading .xls requires xlrd — pip install xlrd"
            ) from exc
        if isinstance(path_or_bytes, bytes):
            workbook = xlrd.open_workbook(file_contents=path_or_bytes)
        else:
            workbook = xlrd.open_workbook(str(path_or_bytes))
        sheet = workbook.sheet_by_index(0)
        rows = [
            [_cell_to_text(sheet.cell_value(r, c)) for c in range(sheet.ncols)]
            for r in range(sheet.nrows)
        ]
        return sheet.name, rows

    source = BytesIO(path_or_bytes) if isinstance(path_or_bytes, bytes) else path_or_bytes
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = [
            [_cell_to_text(value) for value in values]
            for values in sheet.iter_rows(values_only=True)
        ]
        return sheet.title, rows
    finally:
        workbook.close()


def _detect_header_row_index(rows: list[list[str]]) -> int:
    """Pick the row that looks most like a header within the first 20 rows."""
    candidates = rows[:20]
    best_index = 0
    best_score = -1
    for index, row in enumerate(candidates):
        values = [v.strip() for v in row if v.strip()]
        if not values:
            continue
        non_empty = len(values)
        next_non_empty = (
            len([v.strip() for v in rows[index + 1] if v.strip()])
            if index + 1 < len(rows) else 0
        )
        known_hits = sum(1 for v in values if v in KNOWN_HEADER_WORDS)
        score = non_empty * 10 + known_hits * 8 + min(next_non_empty, non_empty)
        # A row with a single cell followed by a wider row is usually a title
        if non_empty == 1 and next_non_empty > 1:
            score -= 20
        if score > best_score:
            best_index = index
            best_score = score
    return best_index


# ── Preview + config ────────────────────────────────────────────────

def parse_excel_preview(
    path_or_bytes: str | Path | bytes, sample_limit: int = 3
) -> dict[str, Any]:
    sheet_name, rows = _read_excel_rows(path_or_bytes)
    if not rows:
        raise ValueError("Excel is empty")

    header_row_index = _detect_header_row_index(rows)
    columns = rows[header_row_index]
    if not any(columns):
        raise ValueError("Excel header row is empty")

    sample_rows: list[dict[str, str]] = []
    row_count = 0
    for values in rows[header_row_index + 1 :]:
        if not values or not any(v.strip() for v in values):
            continue
        row_count += 1
        if len(sample_rows) < sample_limit:
            sample_rows.append({
                column: values[i] if i < len(values) else ""
                for i, column in enumerate(columns)
                if column
            })

    return {
        "columns": [c for c in columns if c],
        "sample_rows": sample_rows,
        "row_count": row_count,
        "sheet_name": sheet_name,
        "header_row": header_row_index + 1,
    }


def guess_excel_config(columns: list[str]) -> dict[str, Any]:
    def first_match(candidates: list[str], fallback: str = "") -> str:
        lowered = [(c, c.lower()) for c in columns]
        for keyword in candidates:
            kw = keyword.lower()
            for c, cl in lowered:
                if kw in cl:
                    return c
        return fallback

    title_field = first_match(
        ["标题", "名称", "政策名称", "title"], columns[0] if columns else ""
    )
    content_field = first_match(
        ["政策原文", "正文", "内容", "content", "text"], columns[-1] if columns else ""
    )
    filter_field = first_match(["级别", "地区", "类型", "分类", "level"], "")
    source_field = first_match(
        ["政策发文字号", "发文字号", "文号", "来源", "source"], ""
    )
    ignore_fields = [c for c in columns if c in DEFAULT_IGNORE_FIELDS]
    display_fields = [f for f in [title_field, filter_field, source_field] if f]

    return {
        "document_mode": "row_as_document",
        "title_field": title_field,
        "content_fields": [content_field] if content_field else [],
        "filter_fields": [filter_field] if filter_field else [],
        "source_fields": [source_field] if source_field else [],
        "display_fields": display_fields,
        "ignore_fields": ignore_fields,
        "chunking": {
            "enabled": True,
            "strategy": "fixed_overlap",
            "fallback_chunk_size": DEFAULT_EXCEL_CHUNK_SIZE,
            "overlap": DEFAULT_EXCEL_CHUNK_OVERLAP,
        },
    }


def normalize_excel_config(
    config: dict[str, Any], columns: list[str]
) -> dict[str, Any]:
    column_set = set(columns)

    def one(field_name: str) -> str:
        value = str(config.get(field_name) or "").strip()
        if value not in column_set:
            raise ValueError(f"{field_name} not in headers: {value}")
        return value

    def many(field_name: str) -> list[str]:
        raw = config.get(field_name) or []
        if isinstance(raw, str):
            raw = [item.strip() for item in re.split(r"[,，\n]", raw) if item.strip()]
        values: list[str] = []
        for value in raw:
            v = str(value or "").strip()
            if not v:
                continue
            if v not in column_set:
                raise ValueError(f"{field_name} unknown header: {v}")
            if v not in values:
                values.append(v)
        return values

    title_field = one("title_field")
    content_fields = many("content_fields")
    if not content_fields:
        raise ValueError("at least one content field is required")

    chunking = config.get("chunking") or {}
    chunk_size = int(chunking.get("fallback_chunk_size") or DEFAULT_EXCEL_CHUNK_SIZE)
    overlap = int(chunking.get("overlap") or DEFAULT_EXCEL_CHUNK_OVERLAP)
    chunk_size = max(200, min(5000, chunk_size))
    overlap = max(0, min(chunk_size - 1, overlap))

    return {
        "document_mode": "row_as_document",
        "title_field": title_field,
        "content_fields": content_fields,
        "filter_fields": many("filter_fields"),
        "source_fields": many("source_fields"),
        "display_fields": many("display_fields") or [title_field],
        "ignore_fields": many("ignore_fields"),
        "chunking": {
            "enabled": True,
            "strategy": "fixed_overlap",
            "fallback_chunk_size": chunk_size,
            "overlap": overlap,
        },
    }


# ── Chunking ─────────────────────────────────────────────────────────

def fixed_overlap_chunks(text: str, chunk_size: int, overlap: int) -> list[str]:
    normalized = "\n".join(
        line.strip() for line in str(text or "").splitlines() if line.strip()
    )
    if not normalized:
        return []
    if len(normalized) <= chunk_size:
        return [normalized]

    chunks: list[str] = []
    start = 0
    step = max(1, chunk_size - overlap)
    while start < len(normalized):
        chunk = normalized[start : start + chunk_size].strip()
        if chunk:
            chunks.append(chunk)
        if start + chunk_size >= len(normalized):
            break
        start += step
    return chunks


def _join_labeled_fields(row: dict[str, str], fields: list[str]) -> str:
    parts = []
    for field in fields:
        value = row.get(field, "").strip()
        if value:
            parts.append(f"{field}：{value}")
    return "\n".join(parts)


# ── Ingestion ────────────────────────────────────────────────────────

async def ingest_excel(
    *,
    kb_id: str,
    doc_id: str,
    file_path: str,
    creds: RequestCredentials,
    config: dict[str, Any] | None = None,
    progress: Any = None,  # ProgressCb but optional for tests
) -> dict[str, Any]:
    """Run Excel ingestion. If `config` is None, auto-guess from headers.

    Pipeline:
      1) Parse rows + detect header
      2) Build policies (row → metadata + content + chunks)
      3) Write policies + chunks to rag.excel_policy(_chunk) (atomic)
      4) Embed search_text per chunk
    """
    if not Path(file_path).exists():
        raise FileNotFoundError(file_path)
    preview = parse_excel_preview(file_path)
    columns = preview["columns"]
    if config is None:
        config = guess_excel_config(columns)
    normalized = normalize_excel_config(config, columns)

    _emit(progress, 10, "parsing Excel rows")
    sheet_name, raw_rows = _read_excel_rows(file_path)
    header_row_index = _detect_header_row_index(raw_rows)
    header_values = raw_rows[header_row_index] if raw_rows else []
    header_cols = [_cell_to_text(v) for v in header_values]
    chunking = normalized["chunking"]
    chunk_size = int(chunking["fallback_chunk_size"])
    overlap = int(chunking["overlap"])
    file_name = Path(file_path).name

    policies: list[dict[str, Any]] = []
    for source_row, values in enumerate(
        raw_rows[header_row_index + 1 :], start=header_row_index + 2
    ):
        row = {
            column: values[i] if i < len(values) else ""
            for i, column in enumerate(header_cols)
            if column
        }
        if not any(row.values()):
            continue

        title = row.get(normalized["title_field"], "").strip() or f"Row {source_row}"
        content = "\n".join(
            row.get(field, "").strip()
            for field in normalized["content_fields"]
            if row.get(field, "").strip()
        )
        if not content:
            continue

        metadata = {
            field: row.get(field, "").strip()
            for field in [*normalized["filter_fields"], *normalized["source_fields"]]
            if row.get(field, "").strip()
        }
        metadata.update({
            "title": title,
            "source_file": file_name,
            "source_sheet": sheet_name,
            "source_row": source_row,
        })
        metadata_text = _join_labeled_fields(
            row,
            [*normalized["filter_fields"], *normalized["source_fields"]],
        )

        chunks = []
        for chunk_index, chunk_text in enumerate(
            fixed_overlap_chunks(content, chunk_size=chunk_size, overlap=overlap),
            start=1,
        ):
            search_text = "\n".join(
                part for part in [
                    f"标题：{title}",
                    metadata_text,
                    f"正文片段：{chunk_text}",
                ] if part
            )
            chunks.append({
                "chunk_index": chunk_index,
                "chunk_text": chunk_text,
                "search_text": search_text,
                "search_text_tokens": tokenize_for_index(search_text),
            })

        policies.append({
            "source_row": source_row,
            "title": title,
            "content": content,
            "metadata": metadata,
            "chunks": chunks,
        })

    _emit(progress, 40, f"writing {len(policies)} policies")
    indexed_count = await replace_excel_policies(
        kb_id=kb_id, doc_id=doc_id, policies=policies,
    )
    await save_excel_config(
        kb_id=kb_id, doc_id=doc_id, config=normalized, row_count=indexed_count,
    )

    # Embed chunks
    vector_count = 0
    if creds.embedding_api_key:
        _emit(progress, 60, "embedding chunks")
        pending = await list_excel_chunks_for_embedding(kb_id=kb_id, doc_id=doc_id)
        if pending:
            texts = [c["search_text"] for c in pending]
            # Cap at 5000 chunks per ingest to avoid runaway cost
            if len(texts) > 5000:
                logger.warning("Excel doc has %d chunks; capping embedding at 5000", len(texts))
                texts = texts[:5000]
                pending = pending[:5000]
            embeddings = encode_texts(creds, texts)
            for chunk, emb in zip(pending, embeddings):
                await save_excel_chunk_vector(
                    chunk_id=int(chunk["chunk_id"]), embedding=emb,
                )
            vector_count = len(embeddings)
    else:
        logger.warning("no embedding key — Excel chunks have FTS but no vectors")

    _emit(progress, 100, f"done — {indexed_count} policies / {vector_count} vectors")
    return {
        "indexed_count": indexed_count,
        "vector_count": vector_count,
        "config": normalized,
    }


def _emit(progress: Any, pct: float, msg: str) -> None:
    if callable(progress):
        progress(float(pct), msg)
    else:
        logger.info("[excel] %s%% — %s", pct, msg)
